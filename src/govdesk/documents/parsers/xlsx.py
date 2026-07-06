# SPDX-FileCopyrightText: 2026 GovDesk-Mitwirkende
#
# SPDX-License-Identifier: EUPL-1.2

"""XLSX-Parser (openpyxl): jede Tabelle wird zur Sektion, Zeilen zu Textzeilen."""

import io

from openpyxl import load_workbook

from govdesk.documents.parsers.base import Block, ParsedDocument


def _cell(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


class XlsxParser:
    def parse(self, data: bytes) -> ParsedDocument:
        # read_only + data_only: schnell und ohne Formeln (nur berechnete Werte).
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        blocks: list[Block] = []
        for index, sheet in enumerate(workbook.worksheets, start=1):
            blocks.append(Block(text=sheet.title, heading_level=2, page_no=index))
            for row in sheet.iter_rows(values_only=True):
                cells = [_cell(v) for v in row]
                line = " | ".join(c for c in cells if c)
                if line:
                    blocks.append(Block(text=line, page_no=index))
        workbook.close()
        return ParsedDocument(blocks=blocks)
